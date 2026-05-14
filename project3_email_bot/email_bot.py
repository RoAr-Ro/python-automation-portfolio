import os
import requests
import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment

# Configuration
EMAIL = os.environ.get("EMAIL")
PASSWORD = os.environ.get("PASSWORD")
RECIPIENT = os.environ.get("RECIPIENT")

# Fetch data
response = requests.get("https://jsonplaceholder.typicode.com/users")
users = response.json()

# Create Excel report
wb = openpyxl.Workbook()
sheet = wb.active
sheet.title = "Users"

# Headers with professional formatting
headers = ["Name", "Email", "City", "Phone"]
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(fill_type="solid", fgColor="2F75B6")
header_alignment = Alignment(horizontal="center")

for col, header in enumerate(headers, 1):
    cell = sheet.cell(row=1, column=col)
    cell.value = header
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment

for user in users:
    sheet.append([
        user['name'],
        user['email'],
        user['address']['city'],
        user['phone']
    ])

wb.save("users_report.xlsx")

# Email setup
today = datetime.now()
user_count = len(users)

message = MIMEMultipart()
message["From"] = EMAIL
message["To"] = RECIPIENT
message["Subject"] = f"Users Report - {today.day:02d}/{today.month:02d}/{today.year}"

body = f"Processed {user_count} users successfully. Please find the report attached."
message.attach(MIMEText(body, "plain"))

# Attach file
with open("users_report.xlsx", "rb") as attachment:
    part = MIMEBase("application", "octet-stream")
    part.set_payload(attachment.read())

encoders.encode_base64(part)
part.add_header(
    "Content-Disposition",
    "attachment; filename=users_report.xlsx"
)
message.attach(part)

# Send email
with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
    server.login(EMAIL, PASSWORD)
    server.sendmail(EMAIL, RECIPIENT, message.as_string())

print(f"✅ Email sent successfully with {user_count} users report!")