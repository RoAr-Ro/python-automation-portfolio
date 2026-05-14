import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import requests

try:
    # Fetch data
    response = requests.get("https://jsonplaceholder.typicode.com/users")
    response.raise_for_status()
    users = response.json()

    # Create workbook
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Users"

    # Headers with professional formatting
    headers = ["Name", "Email", "City"]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="2F75B6")
    header_alignment = Alignment(horizontal="center")

    for col, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for user in users:
        sheet.append([
            user['name'],
            user['email'],
            user['address']['city']
        ])

    # Auto column width
    for column in sheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        sheet.column_dimensions[column_letter].width = max_length + 4

    wb.save("users_report.xlsx")
    print(f"Report created with {len(users)} users!")

except requests.exceptions.ConnectionError:
    print("❌ No internet connection")
except requests.exceptions.HTTPError as e:
    print(f"❌ HTTP Error: {e}")
except Exception as e:
    print(f"❌ Unexpected error: {e}")